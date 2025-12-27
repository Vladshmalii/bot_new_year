import openpyxl
from typing import Dict, List, Any
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from datetime import datetime
import json

from database import Character, Mob, Location, Item, NoteTemplate


async def import_excel_data(file_path: str, db: AsyncSession) -> Dict[str, Any]:
    """
    Import data from Excel file.
    Expected sheets: characters, mobs, locations, items, notes_templates
    """
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    result = {
        "characters": {"created": 0, "updated": 0, "errors": []},
        "mobs": {"created": 0, "updated": 0, "errors": []},
        "locations": {"created": 0, "updated": 0, "errors": []},
        "items": {"created": 0, "updated": 0, "errors": []},
        "notes_templates": {"created": 0, "updated": 0, "errors": []}
    }
    
    # Import characters
    if "characters" in workbook.sheetnames:
        sheet = workbook["characters"]
        headers = [cell.value for cell in sheet[1]]
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            try:
                row_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
                
                if not row_data.get("name"):
                    continue
                
                # Check if character exists (by id or name)
                char_id = row_data.get("id")
                if char_id:
                    query = select(Character).where(Character.id == char_id)
                    existing = (await db.execute(query)).scalar_one_or_none()
                else:
                    query = select(Character).where(Character.name == row_data["name"])
                    existing = (await db.execute(query)).scalar_one_or_none()
                
                # Parse stats
                stats = {}
                if row_data.get("stats"):
                    try:
                        if isinstance(row_data["stats"], str):
                            stats = json.loads(row_data["stats"])
                        else:
                            stats = row_data["stats"]
                    except:
                        # Try parsing individual columns
                        for key in ["str", "dex", "int", "con", "wis", "cha"]:
                            if key in row_data and row_data[key] is not None:
                                stats[key] = int(row_data[key])
                
                # Parse abilities
                abilities = []
                if row_data.get("abilities"):
                    if isinstance(row_data["abilities"], str):
                        abilities = [a.strip() for a in row_data["abilities"].split(",")]
                    elif isinstance(row_data["abilities"], list):
                        abilities = row_data["abilities"]
                
                if existing:
                    # Update existing
                    existing.name = row_data.get("name", existing.name)
                    existing.age = row_data.get("age") or existing.age
                    existing.description = row_data.get("description") or existing.description
                    existing.backstory = row_data.get("backstory") or existing.backstory
                    existing.hp_max = row_data.get("hp_max") or existing.hp_max
                    existing.hp_current = row_data.get("hp_current") or existing.hp_current or existing.hp_max
                    existing.damage_base = row_data.get("damage_base") or existing.damage_base
                    if stats:
                        existing.stats = stats
                    if abilities:
                        existing.abilities = abilities
                    result["characters"]["updated"] += 1
                else:
                    # Create new
                    character = Character(
                        name=row_data["name"],
                        age=row_data.get("age"),
                        description=row_data.get("description"),
                        backstory=row_data.get("backstory"),
                        hp_max=row_data.get("hp_max", 100),
                        hp_current=row_data.get("hp_current") or row_data.get("hp_max", 100),
                        damage_base=row_data.get("damage_base", 1),
                        stats=stats,
                        abilities=abilities
                    )
                    db.add(character)
                    result["characters"]["created"] += 1
            except Exception as e:
                result["characters"]["errors"].append(f"Row {row_idx}: {str(e)}")
        
        await db.commit()
    
    # Import mobs
    if "mobs" in workbook.sheetnames:
        sheet = workbook["mobs"]
        headers = [cell.value for cell in sheet[1]]
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            try:
                row_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
                
                if not row_data.get("name"):
                    continue
                
                mob_id = row_data.get("id")
                if mob_id:
                    query = select(Mob).where(Mob.id == mob_id)
                    existing = (await db.execute(query)).scalar_one_or_none()
                else:
                    query = select(Mob).where(Mob.name == row_data["name"])
                    existing = (await db.execute(query)).scalar_one_or_none()
                
                if existing:
                    existing.name = row_data.get("name", existing.name)
                    existing.description = row_data.get("description") or existing.description
                    existing.base_hp = row_data.get("base_hp") or existing.base_hp
                    existing.base_damage = row_data.get("base_damage") or existing.base_damage
                    existing.dice_pattern = row_data.get("dice_pattern") or existing.dice_pattern
                    existing.public_description = row_data.get("public_description") or existing.public_description
                    existing.gm_notes = row_data.get("gm_notes") or existing.gm_notes
                    result["mobs"]["updated"] += 1
                else:
                    mob = Mob(
                        name=row_data["name"],
                        description=row_data.get("description"),
                        base_hp=row_data.get("base_hp", 50),
                        base_damage=row_data.get("base_damage", 1),
                        dice_pattern=row_data.get("dice_pattern"),
                        public_description=row_data.get("public_description"),
                        gm_notes=row_data.get("gm_notes")
                    )
                    db.add(mob)
                    result["mobs"]["created"] += 1
            except Exception as e:
                result["mobs"]["errors"].append(f"Row {row_idx}: {str(e)}")
        
        await db.commit()
    
    # Import locations
    if "locations" in workbook.sheetnames:
        sheet = workbook["locations"]
        headers = [cell.value for cell in sheet[1]]
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            try:
                row_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
                
                if not row_data.get("name"):
                    continue
                
                loc_id = row_data.get("id")
                if loc_id:
                    query = select(Location).where(Location.id == loc_id)
                    existing = (await db.execute(query)).scalar_one_or_none()
                else:
                    query = select(Location).where(Location.name == row_data["name"])
                    existing = (await db.execute(query)).scalar_one_or_none()
                
                # Parse tags
                tags = []
                if row_data.get("tags"):
                    if isinstance(row_data["tags"], str):
                        tags = [t.strip() for t in row_data["tags"].split(",")]
                    elif isinstance(row_data["tags"], list):
                        tags = row_data["tags"]
                
                if existing:
                    existing.name = row_data.get("name", existing.name)
                    existing.description = row_data.get("description") or existing.description
                    if tags:
                        existing.tags = tags
                    result["locations"]["updated"] += 1
                else:
                    location = Location(
                        name=row_data["name"],
                        description=row_data.get("description"),
                        tags=tags
                    )
                    db.add(location)
                    result["locations"]["created"] += 1
            except Exception as e:
                result["locations"]["errors"].append(f"Row {row_idx}: {str(e)}")
        
        await db.commit()
    
    # Import items
    if "items" in workbook.sheetnames:
        sheet = workbook["items"]
        headers = [cell.value for cell in sheet[1]]
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            try:
                row_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
                
                if not row_data.get("name"):
                    continue
                
                item_id = row_data.get("id")
                if item_id:
                    query = select(Item).where(Item.id == item_id)
                    existing = (await db.execute(query)).scalar_one_or_none()
                else:
                    query = select(Item).where(Item.name == row_data["name"])
                    existing = (await db.execute(query)).scalar_one_or_none()
                
                # Parse base_stats
                base_stats = {}
                if row_data.get("base_stats"):
                    try:
                        if isinstance(row_data["base_stats"], str):
                            base_stats = json.loads(row_data["base_stats"])
                        else:
                            base_stats = row_data["base_stats"]
                    except:
                        pass
                
                if existing:
                    existing.name = row_data.get("name", existing.name)
                    existing.short_description = row_data.get("short_description") or existing.short_description
                    existing.long_description = row_data.get("long_description") or existing.long_description
                    if base_stats:
                        existing.base_stats = base_stats
                    existing.rarity = row_data.get("rarity") or existing.rarity
                    existing.charges = row_data.get("charges") or existing.charges
                    existing.cooldown = row_data.get("cooldown") or existing.cooldown
                    result["items"]["updated"] += 1
                else:
                    item = Item(
                        name=row_data["name"],
                        short_description=row_data.get("short_description"),
                        long_description=row_data.get("long_description"),
                        base_stats=base_stats,
                        rarity=row_data.get("rarity"),
                        charges=row_data.get("charges", 0),
                        cooldown=row_data.get("cooldown", 0)
                    )
                    db.add(item)
                    result["items"]["created"] += 1
            except Exception as e:
                result["items"]["errors"].append(f"Row {row_idx}: {str(e)}")
        
        await db.commit()
    
    # Import note templates
    if "notes_templates" in workbook.sheetnames:
        sheet = workbook["notes_templates"]
        headers = [cell.value for cell in sheet[1]]
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            try:
                row_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
                
                if not row_data.get("text"):
                    continue
                
                template_id = row_data.get("id")
                if template_id:
                    query = select(NoteTemplate).where(NoteTemplate.id == template_id)
                    existing = (await db.execute(query)).scalar_one_or_none()
                else:
                    query = select(NoteTemplate).where(NoteTemplate.text == row_data["text"])
                    existing = (await db.execute(query)).scalar_one_or_none()
                
                if existing:
                    existing.text = row_data.get("text", existing.text)
                    existing.visibility = row_data.get("visibility", existing.visibility)
                    result["notes_templates"]["updated"] += 1
                else:
                    template = NoteTemplate(
                        text=row_data["text"],
                        visibility=row_data.get("visibility", "decide_yourself")
                    )
                    db.add(template)
                    result["notes_templates"]["created"] += 1
            except Exception as e:
                result["notes_templates"]["errors"].append(f"Row {row_idx}: {str(e)}")
        
        await db.commit()
    
    return result

